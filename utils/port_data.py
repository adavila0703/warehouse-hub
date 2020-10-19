from openpyxl import load_workbook
import sqlite3
from datetime import datetime


def lt_check(time1, time2):
    """Checks time between two dates"""
    try:
        tm1 = datetime(int(time1.split()[0].split('-')[0]), int(time1.split()[0].split('-')[1]),
                       int(time1.split()[0].split('-')[2]), int(time1.split()[1].split(':')[0]),
                       int(time1.split()[1].split(':')[1]), int(time1.split()[1].split(':')[2].split('.')[0]))

        tm2 = datetime(int(time2.split()[0].split('-')[0]), int(time2.split()[0].split('-')[1]),
                       int(time2.split()[0].split('-')[2]), int(time2.split()[1].split(':')[0]),
                       int(time2.split()[1].split(':')[1]), int(time2.split()[1].split(':')[2].split('.')[0]))
        return str(tm2 - tm1)
    except IndexError:
        return ''


def port_micro():
    """Function meant to port all micro data to the new database"""
    lb = load_workbook('micro.xlsx')

    ws = lb.active
    sheet_range = lb['Sheet1']

    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    obj = []

    lettercount = 0
    count = 1

    connection = sqlite3.connect('data.db')
    cursor = connection.cursor()

    while True:
        if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
            while True:
                if lettercount <= 11:
                    if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
                        obj.append(str(sheet_range[f'{letters[lettercount]}{count}'].value))
                    else:
                        obj.append('')
                    lettercount += 1
                else:
                    lettercount = 0
                    count += 1
                    rtos = lt_check(obj[5], obj[7])
                    stoc = lt_check(obj[7], obj[10])
                    query = 'INSERT INTO mag_table(date_created, employee, item, serial_num, ins_type, rec_date, ' \
                            'start_date, accessories, appearance, functions, cleaning, complete, notes) ' \
                            'VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)'
                    cursor.execute(query,
                                   (obj[0], obj[1], obj[2], obj[3], obj[4], obj[5], obj[6], obj[7], obj[8], obj[9],
                                    obj[10], rtos, stoc))
                    connection.commit()
                    obj = []
                    break

        else:
            break


def port_marking():
    """Function meant to port all marking data to the new database"""
    lb = load_workbook('marking.xlsx')

    ws = lb.active
    sheet_range = lb['Sheet1']

    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
    obj = []

    lettercount = 0
    count = 1

    connection = sqlite3.connect('data.db')
    cursor = connection.cursor()

    while True:
        if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
            while True:
                if lettercount <= 11:
                    if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
                        obj.append(str(sheet_range[f'{letters[lettercount]}{count}'].value))
                    else:
                        obj.append('')
                    lettercount += 1
                else:
                    lettercount = 0
                    count += 1
                    rtos = lt_check(obj[5], obj[7])
                    stoc = lt_check(obj[7], obj[10])
                    query = 'INSERT INTO mag_table(date_created, employee, item, serial_num, ins_type, rec_date, ' \
                            'start_date, accessories, appearance, functions, cleaning, complete, notes) ' \
                            'VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)'
                    cursor.execute(query,
                                   (obj[0], obj[1], obj[2], obj[3], obj[4], obj[5], obj[6], obj[7], obj[8], obj[9],
                                    obj[10], rtos, stoc))
                    connection.commit()
                    obj = []
                    break

        else:
            break


def port_printing():
    lb = load_workbook('pct.xlsx')

    ws = lb.active
    sheet_range = lb['Sheet1']

    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
               'V', 'W']
    obj = []

    lettercount = 0
    count = 1

    connection = sqlite3.connect('data.db')
    cursor = connection.cursor()

    while True:
        if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
            while True:
                if lettercount <= 22:
                    if sheet_range[f'{letters[lettercount]}{count}'].value is not None:
                        obj.append(str(sheet_range[f'{letters[lettercount]}{count}'].value))
                    else:
                        obj.append('')
                    lettercount += 1
                else:
                    lettercount = 0
                    count += 1
                    rtos = lt_check(obj[5], obj[7])
                    stoc = lt_check(obj[7], obj[10])
                    query = 'INSERT INTO mag_table(date_created, employee, item, serial_num, ins_type, rec_date, ' \
                            'start_date, accessories, appearance, functions, cleaning, complete, notes) ' \
                            'VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)'
                    cursor.execute(query,
                                   (obj[0], obj[1], obj[2], obj[3], obj[4], obj[5], obj[6], obj[7], obj[8], obj[9],
                                    obj[10], rtos, stoc))
                    connection.commit()
                    obj = []
                    break

        else:
            break


if __name__ == '__main__':
    port_micro()
    port_marking()
    port_printing()
